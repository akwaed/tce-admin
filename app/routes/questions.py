"""
Question Bank Routes
View, edit, add, remove questions with approval workflow for department admins
"""
from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, Response
from flask_login import login_required, current_user
from functools import wraps
import pandas as pd
import os
import io
import json
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
from app.models import db
from app.models.question import QBAuditLog
from app.services.backup_service import get_backup_service

questions_bp = Blueprint('questions', __name__)

# Data storage paths
DATASOURCES_PATH = './datasources'
QB_FILENAME = 'QB.xlsx'
QM_FILENAME = 'QM.xlsx'
QM_SHEET_NAME = 'Question Bank Mappings'
PENDING_CHANGES_FILE = 'pending_changes.json'
AUDIT_LOG_FILE = 'qb_audit_log.json'


def qb_access_required(f):
    """Decorator to check Question Bank access"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.can_manage_qb():
            flash('You do not have access to the Question Bank.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def api_login_required(f):
    """Decorator for API endpoints that returns JSON errors instead of redirects for auth failures"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({
                'success': False,
                'error': 'Session expired. Please refresh the page and log in again.'
            }), 401
        return f(*args, **kwargs)
    return decorated_function


def api_qb_access_required(f):
    """Decorator for API endpoints to check Question Bank access (returns JSON errors instead of redirects)"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({
                'error': 'Session expired. Please refresh the page and log in again.',
                'course_questions': [],
                'instructor_questions': [],
                'available_placeholders': []
            }), 401
        if not current_user.can_manage_qb():
            return jsonify({
                'error': 'You do not have access to the Question Bank.',
                'course_questions': [],
                'instructor_questions': [],
                'available_placeholders': []
            }), 403
        return f(*args, **kwargs)
    return decorated_function


def log_audit(action, user, details):
    """Log audit entry for question bank changes"""
    audit_path = os.path.join(DATASOURCES_PATH, AUDIT_LOG_FILE)
    
    try:
        os.makedirs(DATASOURCES_PATH, exist_ok=True)
        if os.path.exists(audit_path):
            with open(audit_path, 'r') as f:
                audit_log = json.load(f)
        else:
            audit_log = []
        
        audit_log.append({
            'timestamp': datetime.now().isoformat(),
            'action': action,
            'user': user.linkblue,
            'user_role': user.role,
            'college': getattr(user, 'college_code', None),
            'department': getattr(user, 'department_id', None),
            'details': details
        })
        
        with open(audit_path, 'w') as f:
            json.dump(audit_log, f, indent=2)
    except Exception as e:
        print(f"Audit log error: {e}")


class PendingChangesManager:
    """Manage pending changes from department admins requiring approval"""
    
    def __init__(self, datafiles_path=DATASOURCES_PATH):
        self.filepath = os.path.join(datafiles_path, PENDING_CHANGES_FILE)
        self.changes = self._load()
    
    def _load(self):
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, 'r') as f:
                    return json.load(f)
            except:
                return []
        return []
    
    def _save(self):
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        with open(self.filepath, 'w') as f:
            json.dump(self.changes, f, indent=2)
    
    def add_change(self, change_type, unit_type, unit_id, placeholder, question_id,
                   submitted_by, college_code=None, old_value=None, new_text=None,
                   old_type=None, new_type=None):
        """Add a pending change request"""
        change = {
            'id': max([c['id'] for c in self.changes], default=0) + 1,
            'type': change_type,  # 'add', 'remove', 'edit'
            'unit_type': unit_type,
            'unit_id': unit_id,
            'placeholder': placeholder,
            'question_id': question_id,
            'old_value': old_value,
            'new_text': new_text,
            'old_type': old_type,
            'new_type': new_type,
            'college_code': college_code,
            'submitted_by': submitted_by,
            'submitted_at': datetime.now().isoformat(),
            'status': 'pending'
        }
        self.changes.append(change)
        self._save()
        return change['id']
    
    def get_pending(self, college_code=None):
        """Get pending changes, optionally filtered by college"""
        pending = [c for c in self.changes if c['status'] == 'pending']
        if college_code:
            pending = [c for c in pending if c.get('college_code') == college_code]
        return pending

    def get_pending_for_submitter(self, submitted_by):
        """Get pending changes submitted by a specific user"""
        return [
            c for c in self.changes
            if c['status'] == 'pending' and c.get('submitted_by') == submitted_by
        ]
    
    def approve(self, change_id, approved_by):
        """Approve a pending change"""
        for change in self.changes:
            if change['id'] == change_id and change['status'] == 'pending':
                change['status'] = 'approved'
                change['approved_by'] = approved_by
                change['approved_at'] = datetime.now().isoformat()
                self._save()
                return change
        return None
    
    def reject(self, change_id, rejected_by, reason=''):
        """Reject a pending change"""
        for change in self.changes:
            if change['id'] == change_id and change['status'] == 'pending':
                change['status'] = 'rejected'
                change['rejected_by'] = rejected_by
                change['rejected_at'] = datetime.now().isoformat()
                change['rejection_reason'] = reason
                self._save()
                return change
        return None
    
    def get_by_id(self, change_id):
        for change in self.changes:
            if change['id'] == change_id:
                return change
        return None


class QuestionBankService:
    """Service class for Question Bank operations with caching"""

    def __init__(self, datafiles_path=DATASOURCES_PATH):
        self.datafiles_path = datafiles_path
        self.questions = {}
        self.question_types = {}
        self.question_mapping = defaultdict(dict)
        self.hierarchy = {}
        self.terms = []
        self._placeholder_names = []
        self._uuid_row = []
        self._unit_to_college = {}
        self._unit_metadata = {}
        self._section_key_to_crs_section = {}
        # Cache tracking
        self._cache_timestamps = {}
        self._hierarchy_html_cache = {}
        self._last_admin_scope = None
        self._last_filter_terms = None

    def _get_file_mtime(self, filename):
        """Get modification time of a file, or 0 if not exists"""
        filepath = os.path.join(self.datafiles_path, filename)
        if os.path.exists(filepath):
            return os.path.getmtime(filepath)
        return 0

    def _is_cache_valid(self, cache_key, filename):
        """Check if cache is still valid based on file modification time"""
        current_mtime = self._get_file_mtime(filename)
        cached_mtime = self._cache_timestamps.get(cache_key, 0)
        return current_mtime > 0 and current_mtime == cached_mtime

    def _update_cache_timestamp(self, cache_key, filename):
        """Update cache timestamp after loading"""
        self._cache_timestamps[cache_key] = self._get_file_mtime(filename)
    
    def load_courses(self, filter_terms=None, admin_scope=None):
        """Load course hierarchy from Courses.csv with caching"""
        courses_file = os.path.join(self.datafiles_path, 'Courses.csv')
        if not os.path.exists(courses_file):
            return {}

        # Check if we can use cached data (same scope and terms, file unchanged)
        scope_key = str(admin_scope) if admin_scope else 'all'
        terms_key = str(sorted(filter_terms)) if filter_terms else 'all'
        cache_valid = (
            self._is_cache_valid('courses', 'Courses.csv') and
            self._last_admin_scope == scope_key and
            self._last_filter_terms == terms_key and
            self.hierarchy
        )

        if cache_valid:
            return self.hierarchy

        df = pd.read_csv(courses_file, low_memory=False)
        self._last_admin_scope = scope_key
        self._last_filter_terms = terms_key
        
        # Get available terms
        if 'ACADEMIC_TERM' in df.columns:
            self.terms = sorted(df['ACADEMIC_TERM'].dropna().unique().tolist())
            if filter_terms:
                df = df[df['ACADEMIC_TERM'].isin(filter_terms)]
        
        # Build unit to college mapping (before filtering)
        full_df = pd.read_csv(courses_file, low_memory=False)
        self._unit_to_college = {}
        self._unit_metadata = {}
        self._section_key_to_crs_section = {}
        for _, row in full_df.iterrows():
            college_code = str(row.get('CLASS_COLLEGE_SHORT', '')).strip()
            college_name = str(row.get('CLASS_COLLEGE', '')).strip()
            dept_id = str(row.get('CLASS_DEPARTMENT_ID', '')).strip()
            dept_name = str(row.get('CLASS_DEPARTMENT', '')).strip()
            class_code = str(row.get('CLASS', '')).strip()
            section_key = str(row.get('SECTION_KEY', '')).strip()
            crs_section = str(row.get('CRS_SECTION', '')).strip()
            section_title = str(row.get('SECTION_TITLE', '')).strip()

            metadata = {
                'college_code': college_code if college_code != 'nan' else '',
                'college_name': college_name if college_name != 'nan' else '',
                'department_id': dept_id if dept_id != 'nan' else '',
                'department_name': dept_name if dept_name != 'nan' else '',
                'course': class_code if class_code != 'nan' else '',
                'section': crs_section if crs_section != 'nan' else '',
                'section_title': section_title if section_title != 'nan' else '',
            }
            
            if dept_id and dept_id != 'nan':
                self._unit_to_college[dept_id] = college_code
                self._unit_metadata.setdefault(('DEPARTMENT', dept_id), metadata)
            if class_code and class_code != 'nan':
                self._unit_to_college[class_code] = college_code
                self._unit_metadata.setdefault(('COURSE', class_code), metadata)
            if section_key and section_key != 'nan':
                self._unit_to_college[section_key] = college_code
                self._unit_metadata.setdefault(('SECTION', section_key), metadata)
            if crs_section and crs_section != 'nan':
                self._unit_to_college[crs_section] = college_code
                self._unit_metadata.setdefault(('SECTION', crs_section), metadata)
                if section_key and section_key != 'nan':
                    self._section_key_to_crs_section[section_key] = crs_section
        
        # Apply admin scope filtering for display
        if admin_scope:
            if admin_scope.get('college'):
                df = df[df['CLASS_COLLEGE_SHORT'] == admin_scope['college']]
            if admin_scope.get('department'):
                df = df[df['CLASS_DEPARTMENT_ID'].astype(str) == str(admin_scope['department'])]
        
        # Build hierarchy
        hierarchy = {}
        for _, row in df.iterrows():
            college_id = str(row.get('CLASS_COLLEGE_SHORT', '')).strip()
            college_name = str(row.get('CLASS_COLLEGE', '')).strip()
            dept_id = str(row.get('CLASS_DEPARTMENT_ID', '')).strip()
            dept_name = str(row.get('CLASS_DEPARTMENT', '')).strip()
            class_code = str(row.get('CLASS', '')).strip()
            section_key = str(row.get('SECTION_KEY', '')).strip()
            
            if not college_id or college_id == 'nan':
                continue
            
            if college_name not in hierarchy:
                hierarchy[college_name] = {
                    'id': college_id,
                    'type': 'college',
                    'children': {}
                }
            
            if dept_name and dept_name != 'nan':
                if dept_name not in hierarchy[college_name]['children']:
                    hierarchy[college_name]['children'][dept_name] = {
                        'id': dept_id,
                        'type': 'department',
                        'children': {}
                    }
                
                if class_code and class_code != 'nan':
                    if class_code not in hierarchy[college_name]['children'][dept_name]['children']:
                        hierarchy[college_name]['children'][dept_name]['children'][class_code] = {
                            'id': class_code,
                            'type': 'course',
                            'children': {}
                        }
                    
                    if section_key and section_key != 'nan':
                        hierarchy[college_name]['children'][dept_name]['children'][class_code]['children'][section_key] = {
                            'id': section_key,
                            'type': 'section',
                            'children': {}
                        }
        
        self.hierarchy = hierarchy
        self._update_cache_timestamp('courses', 'Courses.csv')
        return hierarchy

    def get_college_for_unit(self, unit_id):
        """Get college code for a unit ID"""
        unit_id = str(unit_id)
        return (
            self._unit_to_college.get(unit_id)
            or self._unit_to_college.get(self._canonical_mapping_unit_id('SECTION', unit_id))
        )

    def _strip_term_from_section_key(self, unit_id):
        """Return CRS_SECTION-style ID when passed a SECTION_KEY-style value."""
        unit_id = str(unit_id or '').strip()
        base, separator, suffix = unit_id.rpartition('-')
        if separator and suffix.isdigit() and suffix.startswith('20') and len(suffix) >= 6:
            return base
        return unit_id

    def _canonical_mapping_unit_id(self, unit_type, unit_id):
        """Normalize mapping IDs to the format expected by the QM file."""
        unit_type = str(unit_type or '').upper()
        unit_id = str(unit_id or '').strip()
        if unit_type == 'SECTION':
            return self._section_key_to_crs_section.get(unit_id) or self._strip_term_from_section_key(unit_id)
        return unit_id

    def _merge_question_mappings(self, target, source):
        """Merge mappings without overwriting existing placeholder assignments."""
        for placeholder, question_id in (source or {}).items():
            if placeholder and question_id and not target.get(placeholder):
                target[placeholder] = question_id
        return target

    def _mapping_for_unit(self, unit_type, unit_id, create=False):
        """Get a mapping bucket, coalescing SECTION_KEY aliases into CRS_SECTION IDs."""
        unit_type = str(unit_type or '').upper()
        raw_unit_id = str(unit_id or '').strip()
        canonical_unit_id = self._canonical_mapping_unit_id(unit_type, raw_unit_id)

        if unit_type not in self.question_mapping:
            if not create:
                return {}
            self.question_mapping[unit_type] = {}

        mapping_level = self.question_mapping[unit_type]
        if raw_unit_id != canonical_unit_id and raw_unit_id in mapping_level:
            raw_mapping = mapping_level.pop(raw_unit_id)
            target = mapping_level.setdefault(canonical_unit_id, {})
            self._merge_question_mappings(target, raw_mapping)

        if canonical_unit_id not in mapping_level:
            if not create:
                return {}
            mapping_level[canonical_unit_id] = {}

        return mapping_level[canonical_unit_id]

    def _coalesce_question_mapping(self):
        """Remove duplicate mapping rows that only differ by SECTION_KEY vs CRS_SECTION."""
        for unit_type in list(self.question_mapping.keys()):
            normalized = {}
            for unit_id, questions in self.question_mapping.get(unit_type, {}).items():
                canonical_unit_id = self._canonical_mapping_unit_id(unit_type, unit_id)
                target = normalized.setdefault(canonical_unit_id, {})
                self._merge_question_mappings(target, questions)
            self.question_mapping[unit_type] = normalized

    def _iter_question_mapping_rows(self):
        """Yield normalized QM rows in mapping-level order."""
        self._coalesce_question_mapping()
        for unit_type in ['DEPARTMENT', 'COURSE', 'SECTION']:
            for unit_id, questions in self.question_mapping.get(unit_type, {}).items():
                yield unit_type, unit_id, questions
    
    def load_question_bank(self, qb_file=None):
        """Load questions from QB.xlsx with caching"""
        if qb_file is None:
            qb_file = os.path.join(self.datafiles_path, QB_FILENAME)

        if not os.path.exists(qb_file):
            return {}

        # Check cache validity
        if self._is_cache_valid('qb', QB_FILENAME) and self.questions:
            return self.questions

        try:
            xlsx = pd.ExcelFile(qb_file)
            
            if 'Question Type Definitions' in xlsx.sheet_names:
                types_df = pd.read_excel(xlsx, sheet_name='Question Type Definitions')
                for _, row in types_df.iterrows():
                    type_id = str(row.get('Question Type Definition Id', ''))
                    if type_id and type_id != 'nan':
                        self.question_types[type_id] = {
                            'id': type_id,
                            'name': str(row.get('Question Type Definition Name', '')),
                            'type': str(row.get('Question Type', '')),
                            'options': [row.get(f'Option {i}') for i in range(1, 11) if pd.notna(row.get(f'Option {i}'))]
                        }
            
            if 'Question Bank Questions' in xlsx.sheet_names:
                questions_df = pd.read_excel(xlsx, sheet_name='Question Bank Questions')
                for _, row in questions_df.iterrows():
                    q_id = str(row.get('Question Id', ''))
                    if q_id and q_id != 'nan':
                        type_def_id = str(row.get('Question Type Definition Id', ''))
                        q_type = self.question_types.get(type_def_id, {}).get('type', 'Unknown')
                        
                        self.questions[q_id] = {
                            'id': q_id,
                            'type_id': type_def_id,
                            'type': q_type,
                            'text': str(row.get('Question Title', '')),
                            'detail': str(row.get('Question Detail', '')) if pd.notna(row.get('Question Detail')) else '',
                            'block_title': str(row.get('Block Title', '')) if pd.notna(row.get('Block Title')) else ''
                        }
        except Exception as e:
            print(f"Error loading question bank: {e}")

        self._update_cache_timestamp('qb', QB_FILENAME)
        return self.questions

    def _get_or_create_type_definition(self, question_type):
        """Get a question type definition ID for the given type, creating one if needed."""
        normalized = str(question_type or '').strip()
        if not normalized:
            return ''

        for type_id, t in self.question_types.items():
            if str(t.get('type', '')).strip().lower() == normalized.lower():
                return type_id

        base_id = 'AUTO_SEL' if normalized.lower() == 'selection' else 'AUTO_COM'
        type_id = base_id
        counter = 1
        while type_id in self.question_types:
            counter += 1
            type_id = f"{base_id}_{counter}"

        self.question_types[type_id] = {
            'id': type_id,
            'name': f'Auto {normalized}',
            'type': normalized,
            'options': []
        }
        return type_id
    
    def load_question_mapping(self, qm_file=None):
        """Load question mappings from QM.xlsx with caching"""
        if qm_file is None:
            qm_file = os.path.join(self.datafiles_path, QM_FILENAME)

        if not os.path.exists(qm_file):
            return {}

        # Check cache validity
        if self._is_cache_valid('qm', QM_FILENAME) and self.question_mapping:
            return dict(self.question_mapping)

        try:
            xlsx = pd.ExcelFile(qm_file)
            df = pd.read_excel(xlsx, sheet_name=xlsx.sheet_names[0], header=None)
            
            self._uuid_row = df.iloc[0].tolist()
            self._placeholder_names = [str(x) if pd.notna(x) else '' for x in df.iloc[1].tolist()]
            
            for idx in range(2, len(df)):
                row = df.iloc[idx]
                mapping_type = str(row.iloc[0]).upper() if pd.notna(row.iloc[0]) else ''
                unit_id = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ''
                
                if not mapping_type or mapping_type in ['NAN', 'TYPE'] or not unit_id:
                    continue
                
                questions = {}
                for col_idx in range(2, len(row)):
                    if col_idx < len(self._placeholder_names):
                        placeholder = self._placeholder_names[col_idx]
                        value = row.iloc[col_idx]
                        if pd.notna(value) and str(value) != 'nan':
                            questions[placeholder] = str(value)
                
                canonical_unit_id = self._canonical_mapping_unit_id(mapping_type, unit_id)
                existing_questions = self.question_mapping[mapping_type].setdefault(canonical_unit_id, {})
                self._merge_question_mappings(existing_questions, questions)
        
        except Exception as e:
            print(f"Error loading question mapping: {e}")

        self._update_cache_timestamp('qm', QM_FILENAME)
        return dict(self.question_mapping)

    def get_available_placeholders(self, unit_type):
        """Get available placeholder columns for a unit type"""
        unit_type = unit_type.upper()
        all_placeholders = self._placeholder_names[2:] if len(self._placeholder_names) > 2 else []
        
        # Return all placeholders - filtering can be done in UI
        return [p for p in all_placeholders if p]
    
    def get_questions_for_unit(self, unit_type, unit_id):
        """Get questions mapped to a specific unit"""
        unit_type = unit_type.upper()
        unit_id = str(unit_id)
        
        mapping = self._mapping_for_unit(unit_type, unit_id)
        
        course_questions = []
        instructor_questions = []
        
        for placeholder, q_id in mapping.items():
            placeholder_lower = placeholder.lower()
            
            q_data = self.questions.get(q_id, {
                'id': q_id,
                'type': 'Selection' if 'sel' in placeholder_lower else 'Comment' if 'com' in placeholder_lower else 'Unknown',
                'text': f'Question ID: {q_id}',
                'type_id': '',
                'detail': ''
            })

            q_type = q_data.get('type', 'Unknown')
            if q_type == 'Unknown':
                q_type = 'Selection' if 'sel' in placeholder_lower else 'Comment' if 'com' in placeholder_lower else 'Unknown'

            question = {
                'id': q_id,
                'placeholder': placeholder,
                'type': q_type,
                'type_id': q_data.get('type_id', ''),
                'text': q_data.get('text', ''),
                'detail': q_data.get('detail', '')
            }
            
            if 'ins_' in placeholder_lower or '_ins_' in placeholder_lower:
                instructor_questions.append(question)
            else:
                course_questions.append(question)
        
        return {
            'course_questions': course_questions,
            'instructor_questions': instructor_questions
        }
    
    def get_available_questions(self, search=''):
        """Search available questions for adding"""
        results = []
        search_lower = search.lower()
        for q_id, q in self.questions.items():
            if search_lower in q_id.lower() or search_lower in q.get('text', '').lower():
                results.append({
                    'id': q_id,
                    'type': q.get('type', 'Unknown'),
                    'text': q.get('text', '')[:100] + ('...' if len(q.get('text', '')) > 100 else '')
                })
        return results[:50]
    
    def add_question_to_unit(self, unit_type, unit_id, placeholder, question_id):
        """Add a question mapping to a unit"""
        unit_type = unit_type.upper()
        unit_id = str(unit_id)
        
        mapping = self._mapping_for_unit(unit_type, unit_id, create=True)
        mapping[placeholder] = question_id
        self._save_question_mapping()
        return True
    
    def remove_question_from_unit(self, unit_type, unit_id, placeholder):
        """Remove a question mapping from a unit"""
        unit_type = unit_type.upper()
        unit_id = str(unit_id)
        
        mapping = self._mapping_for_unit(unit_type, unit_id)
        if placeholder in mapping:
            del mapping[placeholder]
            self._save_question_mapping()
            return True
        return False
    
    def update_question(self, question_id, new_text):
        """Update a question's text"""
        return self.update_question_details(question_id, new_text=new_text, new_type=None)

    def update_question_details(self, question_id, new_text=None, new_type=None):
        """Update question text and/or type"""
        if question_id not in self.questions:
            return False

        if new_text is not None:
            self.questions[question_id]['text'] = new_text

        if new_type:
            type_def_id = self._get_or_create_type_definition(new_type)
            self.questions[question_id]['type'] = new_type
            self.questions[question_id]['type_id'] = type_def_id

        self._save_question_bank()
        return True

    def get_next_placeholder(self, unit_type, unit_id, question_type, is_instructor):
        """
        Auto-assign next available placeholder based on unit type and question characteristics

        Placeholder format:
        - Department: Dept_Crs/Ins_Sel/Com_001-014 (14 selection, 5 comment)
        - Course: Crs_Crs/Ins_Sel/Com_001-014 (14 selection, 5 comment)
        - Section: Sec_Crs/Ins_Sel/Com_001-014 (14 selection, 5 comment)
        """
        unit_type = unit_type.upper()
        unit_id = str(unit_id)

        # Map unit type to prefix
        prefix_map = {'DEPARTMENT': 'Dept', 'COURSE': 'Crs', 'SECTION': 'Sec'}
        unit_prefix = prefix_map.get(unit_type, 'Dept')

        # Build placeholder pattern
        target_type = 'Ins' if is_instructor else 'Crs'
        q_type_abbr = 'Sel' if question_type == 'Selection' else 'Com'
        pattern = f"{unit_prefix}_{target_type}_{q_type_abbr}_"

        # Determine max number based on type
        max_num = 14 if q_type_abbr == 'Sel' else 5

        # Get currently used placeholders for this unit
        current_mapping = self._mapping_for_unit(unit_type, unit_id)
        used_placeholders = set(current_mapping.keys())

        # Find next available placeholder
        for i in range(1, max_num + 1):
            placeholder = f"{pattern}{i:03d}"
            if placeholder not in used_placeholders:
                return placeholder

        return None  # No available slots

    def create_new_question(self, question_text, question_type):
        """
        Create a new question and assign it a unique ID
        Returns the new question ID
        """
        if not self.questions:
            self.load_question_bank()

        # Generate new question ID
        # Find highest existing numeric ID and increment
        max_id = 0
        for q_id in self.questions.keys():
            # Try to extract numeric part from question IDs
            try:
                # Handle IDs like "Q_12345" or just "12345"
                numeric_part = ''.join(filter(str.isdigit, str(q_id)))
                if numeric_part:
                    max_id = max(max_id, int(numeric_part))
            except:
                pass

        # Generate new ID with format Q_XXXXXX
        new_id = f"Q_{max_id + 1:06d}"

        type_def_id = self._get_or_create_type_definition(question_type)

        # Add question to questions dictionary
        self.questions[new_id] = {
            'id': new_id,
            'type_id': type_def_id,
            'type': question_type,
            'text': question_text,
            'detail': '',
            'block_title': ''
        }

        # Save to file
        self._save_question_bank()

        return new_id
    
    def _save_question_bank(self):
        """Save question bank to Excel"""
        output_path = os.path.join(self.datafiles_path, QB_FILENAME)
        os.makedirs(self.datafiles_path, exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            questions_data = [{
                'Question Id': q_id,
                'Question Type Definition Id': q.get('type_id', ''),
                'Question Type Definition Name': '',
                'Question Title': q.get('text', ''),
                'Question Detail': q.get('detail', ''),
                'Block Title': q.get('block_title', '')
            } for q_id, q in self.questions.items()]
            
            if questions_data:
                pd.DataFrame(questions_data).to_excel(writer, sheet_name='Question Bank Questions', index=False)
            
            types_data = []
            for type_id, t in self.question_types.items():
                row = {
                    'Question Type Definition Id': type_id,
                    'Question Type Definition Name': t.get('name', ''),
                    'Question Type': t.get('type', '')
                }
                for i, opt in enumerate(t.get('options', []), 1):
                    row[f'Option {i}'] = opt
                types_data.append(row)
            
            if types_data:
                pd.DataFrame(types_data).to_excel(writer, sheet_name='Question Type Definitions', index=False)
    
    def _save_question_mapping(self):
        """Save question mapping to Excel"""
        output_path = os.path.join(self.datafiles_path, QM_FILENAME)
        os.makedirs(self.datafiles_path, exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            rows = []
            rows.append(self._uuid_row if self._uuid_row else [''] * 117)
            rows.append(self._placeholder_names if self._placeholder_names else ['Type', 'ID'] + [f'Placeholder_{i}' for i in range(115)])
            
            for unit_type, unit_id, questions in self._iter_question_mapping_rows():
                row = [unit_type, unit_id]
                for i in range(2, len(self._placeholder_names) if self._placeholder_names else 117):
                    placeholder = self._placeholder_names[i] if i < len(self._placeholder_names) else ''
                    row.append(questions.get(placeholder, ''))
                rows.append(row)
            
            pd.DataFrame(rows).to_excel(writer, sheet_name=QM_SHEET_NAME, index=False, header=False)
    
    def get_units_with_questions(self):
        """Get set of unit IDs that have questions assigned"""
        self._coalesce_question_mapping()
        units = {'DEPARTMENT': set(), 'COURSE': set(), 'SECTION': set()}
        for unit_type, mappings in self.question_mapping.items():
            for unit_id, questions in mappings.items():
                if any(q for q in questions.values()):
                    units[unit_type].add(str(unit_id))
        return units
    
    def export_question_bank(self):
        """Export question bank to BytesIO"""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            questions_data = [{
                'Question Id': q_id,
                'Question Type Definition Id': q.get('type_id', ''),
                'Question Type Definition Name': '',
                'Question Title': q.get('text', ''),
                'Question Detail': q.get('detail', ''),
                'Block Title': q.get('block_title', '')
            } for q_id, q in self.questions.items()]
            
            if questions_data:
                pd.DataFrame(questions_data).to_excel(writer, sheet_name='Question Bank Questions', index=False)
            
            types_data = []
            for type_id, t in self.question_types.items():
                row = {'Question Type Definition Id': type_id, 'Question Type Definition Name': t.get('name', ''), 'Question Type': t.get('type', '')}
                for i, opt in enumerate(t.get('options', []), 1):
                    row[f'Option {i}'] = opt
                types_data.append(row)
            
            if types_data:
                pd.DataFrame(types_data).to_excel(writer, sheet_name='Question Type Definitions', index=False)
        
        output.seek(0)
        return output
    
    def export_question_mapping(self):
        """Export question mapping to BytesIO"""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            rows = [
                self._uuid_row if self._uuid_row else [''] * 117,
                self._placeholder_names if self._placeholder_names else ['Type', 'ID'] + [f'Placeholder_{i}' for i in range(115)]
            ]
            
            for unit_type, unit_id, questions in self._iter_question_mapping_rows():
                row = [unit_type, unit_id]
                for i in range(2, len(self._placeholder_names) if self._placeholder_names else 117):
                    placeholder = self._placeholder_names[i] if i < len(self._placeholder_names) else ''
                    row.append(questions.get(placeholder, ''))
                rows.append(row)
            
            pd.DataFrame(rows).to_excel(writer, sheet_name=QM_SHEET_NAME, index=False, header=False)
        
        output.seek(0)
        return output

    def _summary_rows(self, college_code=None):
        """Return readable question assignment rows, optionally scoped to one college."""
        rows = []
        self._coalesce_question_mapping()

        for unit_type, unit_id, questions in self._iter_question_mapping_rows():
            canonical_id = self._canonical_mapping_unit_id(unit_type, unit_id)
            metadata = self._unit_metadata.get((unit_type, canonical_id), {})
            unit_college = metadata.get('college_code') or self.get_college_for_unit(canonical_id) or ''

            # Fail closed for scoped exports: unknown units must never leak into a
            # college administrator's workbook.
            if college_code and unit_college != college_code:
                continue

            department = metadata.get('department_name', '')
            course = metadata.get('course', '')
            section = metadata.get('section', '')
            if unit_type == 'DEPARTMENT':
                department = department or canonical_id
                course = ''
                section = ''
            elif unit_type == 'COURSE':
                course = course or canonical_id
                section = ''
            elif unit_type == 'SECTION':
                section = section or canonical_id

            section_title = metadata.get('section_title', '')
            if section and section_title:
                section = f'{section} — {section_title}'

            for placeholder, question_id in questions.items():
                if not question_id:
                    continue

                placeholder_lower = str(placeholder).lower()
                question = self.questions.get(question_id, {})
                question_type = question.get('type', 'Unknown')
                if question_type == 'Unknown':
                    if 'sel' in placeholder_lower:
                        question_type = 'Selection'
                    elif 'com' in placeholder_lower:
                        question_type = 'Comment'

                type_definition = self.question_types.get(question.get('type_id', ''), {})
                options = ' • '.join(
                    str(option) for option in type_definition.get('options', [])
                    if option is not None and str(option).strip()
                )
                notes = []
                if question.get('block_title'):
                    notes.append(f"Block: {question['block_title']}")
                if question.get('detail'):
                    notes.append(str(question['detail']))

                rows.append({
                    'college_code': unit_college,
                    'college_name': metadata.get('college_name', ''),
                    'values': [
                        unit_type.title(),
                        department,
                        course,
                        section,
                        'Instructor' if 'ins_' in placeholder_lower or '_ins_' in placeholder_lower else 'Course',
                        question_type,
                        question.get('text') or f'Question ID: {question_id}',
                        options,
                        '\n'.join(notes),
                        placeholder,
                        question_id,
                    ]
                })

        rows.sort(key=lambda row: (
            row['college_name'] or row['college_code'],
            row['values'][1],
            row['values'][2],
            row['values'][3],
            row['values'][4],
            row['values'][9],
        ))
        return rows

    @staticmethod
    def _summary_sheet_title(college_code, college_name, used_titles):
        """Create a valid, recognizable, unique Excel sheet title."""
        if college_code:
            base = f'{college_code} - {college_name}' if college_name else college_code
        else:
            base = 'Unassigned Units'
        for character in '[]:*?/\\':
            base = base.replace(character, '-')
        base = base.strip() or 'Question Summary'
        base = base[:31]

        title = base
        counter = 2
        while title.lower() in used_titles:
            suffix = f' ({counter})'
            title = f'{base[:31 - len(suffix)]}{suffix}'
            counter += 1
        used_titles.add(title.lower())
        return title

    def export_question_bank_summary(self, college_code=None, exported_at=None):
        """Export an easy-to-read QB summary, grouped into college worksheets."""
        headers = [
            'Unit Level', 'Department', 'Course', 'Section', 'Audience',
            'Question Type', 'Question', 'Response Options', 'Notes',
            'Placeholder', 'Question ID'
        ]
        rows = self._summary_rows(college_code=college_code)
        grouped_rows = defaultdict(list)
        college_names = {}

        for row in rows:
            key = row['college_code'] or ''
            grouped_rows[key].append(row['values'])
            if row['college_name']:
                college_names[key] = row['college_name']

        # A scoped export should still produce a useful, clearly labeled sheet
        # when that college currently has no mapped questions.
        if college_code and college_code not in grouped_rows:
            grouped_rows[college_code] = []
            for college_name, college in self.hierarchy.items():
                if college.get('id') == college_code:
                    college_names[college_code] = college_name
                    break
        elif not grouped_rows:
            grouped_rows[''] = []

        workbook = Workbook()
        workbook.remove(workbook.active)
        used_titles = set()
        exported_at = exported_at or datetime.now()
        header_fill = PatternFill('solid', fgColor='0033A0')
        subheader_fill = PatternFill('solid', fgColor='E9EFFB')
        light_border = Border(bottom=Side(style='thin', color='B7C9E2'))

        ordered_groups = sorted(
            grouped_rows.items(),
            key=lambda item: (college_names.get(item[0], ''), item[0])
        )
        for sheet_index, (group_code, group_rows) in enumerate(ordered_groups, start=1):
            group_name = college_names.get(group_code, '')
            title = self._summary_sheet_title(group_code, group_name, used_titles)
            sheet = workbook.create_sheet(title=title)
            sheet.sheet_view.showGridLines = False

            label = group_name or group_code or 'Unassigned Units'
            if group_code and group_name:
                label = f'{group_name} ({group_code})'
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            sheet['A1'] = f'Question Bank Summary — {label}'
            sheet['A1'].font = Font(name='Aptos Display', size=18, bold=True, color='FFFFFF')
            sheet['A1'].fill = header_fill
            sheet['A1'].alignment = Alignment(vertical='center')
            sheet.row_dimensions[1].height = 32

            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            sheet['A2'] = (
                f'Exported {exported_at:%Y-%m-%d %H:%M} • '
                f'{len(group_rows):,} assigned question(s)'
            )
            sheet['A2'].font = Font(name='Aptos', size=10, italic=True, color='334155')
            sheet['A2'].fill = subheader_fill
            sheet['A2'].alignment = Alignment(vertical='center')
            sheet.row_dimensions[2].height = 22

            for column, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=column, value=header)
                cell.font = Font(name='Aptos', size=10, bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            sheet.row_dimensions[4].height = 28

            if group_rows:
                for row_index, values in enumerate(group_rows, start=5):
                    for column, value in enumerate(values, start=1):
                        cell = sheet.cell(row=row_index, column=column, value=value)
                        cell.font = Font(name='Aptos', size=10, color='172033')
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                        cell.border = light_border
                    sheet.row_dimensions[row_index].height = 45

                table_ref = f'A4:{get_column_letter(len(headers))}{4 + len(group_rows)}'
                table = Table(displayName=f'QuestionSummary{sheet_index}', ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(
                    name='TableStyleMedium2',
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                sheet.add_table(table)
            else:
                sheet.merge_cells(start_row=5, start_column=1, end_row=6, end_column=len(headers))
                sheet['A5'] = 'No assigned questions were found for this college.'
                sheet['A5'].font = Font(name='Aptos', size=11, italic=True, color='475569')
                sheet['A5'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['A5'].fill = PatternFill('solid', fgColor='F8FAFC')

            widths = [13, 24, 15, 28, 12, 15, 60, 34, 36, 20, 18]
            for column, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(column)].width = width

            sheet.freeze_panes = 'A5'
            sheet.auto_filter.ref = f'A4:{get_column_letter(len(headers))}{max(4, 4 + len(group_rows))}'
            sheet.print_title_rows = '1:4'
            sheet.page_setup.orientation = 'landscape'
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_margins = PageMargins(
                left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.2, footer=0.2
            )
            sheet.oddFooter.center.text = 'Page &P of &N'
            sheet.oddFooter.right.text = 'TCE Question Bank Summary'

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output, {
            'row_count': len(rows),
            'sheet_count': len(ordered_groups),
            'sheet_names': [sheet.title for sheet in workbook.worksheets],
        }
    
    def generate_hierarchy_html(self, units_with_questions=None, use_cache=True):
        """Generate HTML for the hierarchy tree with caching"""
        if units_with_questions is None:
            units_with_questions = self.get_units_with_questions()

        # Create a cache key based on hierarchy and units with questions
        cache_key = f"{self._last_admin_scope}_{self._last_filter_terms}_{hash(str(sorted(str(units_with_questions))))}"
        if use_cache and cache_key in self._hierarchy_html_cache:
            return self._hierarchy_html_cache[cache_key]
        
        def has_questions_recursive(node, node_type, node_id):
            type_map = {'college': 'COLLEGE', 'department': 'DEPARTMENT', 'course': 'COURSE', 'section': 'SECTION'}
            mapped_type = type_map.get(node_type, node_type.upper())
            
            mapped_node_id = self._canonical_mapping_unit_id(mapped_type, node_id)
            if mapped_node_id in units_with_questions.get(mapped_type, set()):
                return True
            
            for child_node in node.get('children', {}).values():
                if has_questions_recursive(child_node, child_node['type'], child_node['id']):
                    return True
            return False
        
        def render_node(name, node, level=0):
            node_id = node['id']
            node_type = node['type']
            children = node.get('children', {})
            
            type_map = {'college': 'COLLEGE', 'department': 'DEPARTMENT', 'course': 'COURSE', 'section': 'SECTION'}
            mapped_type = type_map.get(node_type, node_type.upper())
            
            mapped_node_id = self._canonical_mapping_unit_id(mapped_type, node_id)
            has_direct = mapped_node_id in units_with_questions.get(mapped_type, set())
            has_child = has_questions_recursive(node, node_type, node_id) and not has_direct
            question_count = len(self.question_mapping.get(mapped_type, {}).get(mapped_node_id, {}))
            
            classes = ['tree-node']
            if has_direct:
                classes.append('has-questions')
            
            toggle_classes = ['tree-toggle']
            if has_child:
                toggle_classes.append('has-child-questions')
            
            # Escape single quotes in name for JavaScript
            escaped_name = name.replace("'", "\\'").replace('"', '\\"')
            
            html = f'<div class="tree-item">'
            html += f'<div class="{" ".join(classes)}" onclick="selectUnit(this, \'{mapped_type}\', \'{node_id}\', \'{escaped_name}\')">'
            
            if children:
                html += f'<span class="{" ".join(toggle_classes)}" onclick="toggleTree(this, event)"><i class="bi bi-chevron-right"></i></span>'
            else:
                html += '<span class="tree-toggle"></span>'
            
            html += f'<span class="tree-label">{name}</span>'
            
            if has_direct and question_count > 0:
                html += f'<span class="tree-badge">{question_count}</span>'
            
            html += '</div>'
            
            if children:
                html += '<div class="tree-children">'
                for child_name in sorted(children.keys()):
                    html += render_node(child_name, children[child_name], level + 1)
                html += '</div>'
            
            html += '</div>'
            return html
        
        html = ''
        for college_name in sorted(self.hierarchy.keys()):
            html += render_node(college_name, self.hierarchy[college_name])

        # Cache the result
        self._hierarchy_html_cache[cache_key] = html
        return html


# Global instances
_qb_service = None
_pending_manager = None

def get_qb_service(force_reload=False):
    global _qb_service
    if _qb_service is None or force_reload:
        _qb_service = QuestionBankService(DATASOURCES_PATH)
    return _qb_service

def get_pending_manager():
    global _pending_manager
    if _pending_manager is None:
        _pending_manager = PendingChangesManager(DATASOURCES_PATH)
    return _pending_manager


# ============== ROUTES ==============

@questions_bp.route('/')
@qb_access_required
def browser():
    """Question Bank Browser main view with caching"""
    selected_terms = request.args.getlist('term')

    admin_scope = None
    if not current_user.is_super_admin():
        admin_scope = {
            'college': current_user.college_code,
            'department': current_user.department_id if current_user.role == 'dept_admin' else None
        }

    # Use cached service - caching is handled within each load method
    qb_service = get_qb_service()
    qb_service.load_courses(filter_terms=selected_terms or None, admin_scope=admin_scope)
    qb_service.load_question_bank()
    qb_service.load_question_mapping()

    units_with_questions = qb_service.get_units_with_questions()
    hierarchy_html = qb_service.generate_hierarchy_html(units_with_questions)
    
    # Get pending count for approval badge
    pending_count = 0
    if current_user.is_super_admin() or (current_user.role == 'college_admin'):
        pending_manager = get_pending_manager()
        if current_user.is_super_admin():
            pending_count = len(pending_manager.get_pending())
        else:
            pending_count = len(pending_manager.get_pending(current_user.college_code))

    my_pending_count = 0
    if current_user.role == 'dept_admin':
        pending_manager = get_pending_manager()
        my_pending_count = len(pending_manager.get_pending_for_submitter(current_user.linkblue))
    
    return render_template('questions/browser.html',
                         hierarchy_html=hierarchy_html,
                         terms=qb_service.terms,
                         selected_terms=selected_terms or qb_service.terms,
                         pending_count=pending_count,
                         my_pending_count=my_pending_count,
                         can_approve=current_user.is_super_admin() or current_user.role == 'college_admin',
                         can_export_summary=current_user.is_super_admin() or current_user.role == 'college_admin',
                         is_super_admin=current_user.is_super_admin(),
                         user_role=current_user.role)


@questions_bp.route('/api/questions/<unit_type>/<path:unit_id>')
@api_qb_access_required
def api_get_questions(unit_type, unit_id):
    """API: Get questions for a unit"""
    try:
        qb_service = get_qb_service()

        # Get admin scope
        admin_scope = None
        if not current_user.is_super_admin():
            admin_scope = {
                'college': current_user.college_code,
                'department': current_user.department_id if current_user.role == 'dept_admin' else None
            }

        # Always ensure all data is loaded (check multiple conditions)
        needs_reload = (
            not qb_service.questions or
            not qb_service.question_mapping or
            not qb_service.hierarchy or
            not qb_service._placeholder_names or
            not qb_service._unit_to_college
        )

        if needs_reload:
            # Reload everything to ensure consistency
            qb_service.load_courses(admin_scope=admin_scope)
            qb_service.load_question_bank()
            qb_service.load_question_mapping()

        questions = qb_service.get_questions_for_unit(unit_type, unit_id)

        # Get available placeholders
        all_placeholders = qb_service.get_available_placeholders(unit_type)
        used_placeholders = list(qb_service._mapping_for_unit(unit_type, unit_id).keys())
        unused_placeholders = [p for p in all_placeholders if p not in used_placeholders]

        # Get pending changes for current user (for visual feedback)
        pending_changes = []
        if current_user.role == 'dept_admin':
            pending_manager = get_pending_manager()
            user_pending = pending_manager.get_pending_for_submitter(current_user.linkblue)
            # Filter to this unit's pending changes
            for change in user_pending:
                if change.get('unit_type', '').upper() == unit_type.upper() and str(change.get('unit_id', '')) == str(unit_id):
                    pending_changes.append(change)
                # Also include edit changes for questions in this unit
                elif change.get('type') == 'edit' and change.get('unit_type') == 'QUESTION':
                    # Check if this question is in the current unit
                    q_id = change.get('question_id')
                    for q in questions.get('course_questions', []) + questions.get('instructor_questions', []):
                        if q.get('id') == q_id:
                            pending_changes.append(change)
                            break

        return jsonify({
            **questions,
            'available_placeholders': unused_placeholders[:30],
            'can_edit': current_user.can_manage_qb(),
            'needs_approval': current_user.role == 'dept_admin',
            'pending_changes': pending_changes
        })
    except Exception as e:
        print(f"Error in api_get_questions: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': str(e),
            'course_questions': [],
            'instructor_questions': [],
            'available_placeholders': [],
            'can_edit': current_user.can_manage_qb(),
            'needs_approval': current_user.role == 'dept_admin',
            'pending_changes': []
        }), 500


@questions_bp.route('/api/questions/search')
@api_qb_access_required
def api_search_questions():
    """API: Search available questions"""
    search = request.args.get('q', '')
    qb_service = get_qb_service()
    
    if not qb_service.questions:
        qb_service.load_question_bank()
    
    return jsonify(qb_service.get_available_questions(search))


@questions_bp.route('/api/question/update', methods=['POST'])
@api_qb_access_required
def api_update_question():
    """API: Update question text"""
    data = request.get_json()
    question_id = data.get('question_id')
    question_text = data.get('question_text')
    question_type = data.get('question_type')

    if not question_id:
        return jsonify({'success': False, 'error': 'Question ID required'})

    qb_service = get_qb_service()
    if question_text is None and not question_type:
        return jsonify({'success': False, 'error': 'No updates provided'})

    # Department admins need approval
    if current_user.role == 'dept_admin':
        pending_manager = get_pending_manager()
        old_text = qb_service.questions.get(question_id, {}).get('text', '')
        old_type = qb_service.questions.get(question_id, {}).get('type', '')
        change_id = pending_manager.add_change(
            change_type='edit',
            unit_type='QUESTION',
            unit_id=question_id,
            placeholder=None,
            question_id=question_id,
            submitted_by=current_user.linkblue,
            college_code=current_user.college_code,
            old_value=old_text,
            new_text=question_text,
            old_type=old_type,
            new_type=question_type
        )
        
        log_audit('question_edit_pending', current_user, {'question_id': question_id, 'change_id': change_id})
        return jsonify({'success': True, 'pending': True, 'message': 'Change submitted for approval'})
    
    # Super admin and college admin - apply immediately
    # Create backup before change
    backup_service = get_backup_service()
    backup = backup_service.create_backup('qb', 'change', current_user,
                                           details={'question_id': question_id, 'action': 'edit'})

    if qb_service.update_question_details(question_id, new_text=question_text, new_type=question_type):
        log_audit('question_edit', current_user, {
            'question_id': question_id,
            'updated_text': question_text is not None,
            'updated_type': bool(question_type)
        })
        # Log to database
        QBAuditLog.log_action('question_edit', current_user,
                              details={'question_id': question_id,
                                       'new_text': question_text[:100] if question_text else None,
                                       'new_type': question_type},
                              backup_id=backup.id if backup else None)
        db.session.commit()
        return jsonify({'success': True})

    return jsonify({'success': False, 'error': 'Question not found'})


@questions_bp.route('/api/question/add', methods=['POST'])
@api_qb_access_required
def api_add_question():
    """API: Add question to unit (existing or create new)"""
    data = request.get_json()
    unit_type = data.get('unit_type')
    unit_id = data.get('unit_id')
    question_id = data.get('question_id')
    question_text = data.get('question_text')  # For creating new questions
    question_type = data.get('question_type')  # 'Selection' or 'Comment'
    is_instructor = data.get('is_instructor', False)  # Course vs Instructor question

    if not all([unit_type, unit_id]):
        return jsonify({'success': False, 'error': 'Missing unit information'})

    qb_service = get_qb_service()
    college_code = qb_service.get_college_for_unit(unit_id) or current_user.college_code

    # Creating new question
    if question_text and question_type:
        # Auto-assign placeholder and create question
        placeholder = qb_service.get_next_placeholder(unit_type, unit_id, question_type, is_instructor)
        if not placeholder:
            return jsonify({'success': False, 'error': 'No available placeholders'})

        # Generate new question ID
        question_id = qb_service.create_new_question(question_text, question_type)

        log_audit('question_create', current_user, {
            'question_id': question_id, 'text': question_text, 'type': question_type
        })

    # Adding existing question
    elif question_id:
        # Auto-assign placeholder based on question characteristics
        q = qb_service.questions.get(question_id, {})
        q_type = q.get('type', 'Selection')
        # Infer if instructor question from existing mappings
        is_ins = any(
            'ins_' in p.lower()
            for p, qid in qb_service._mapping_for_unit(unit_type, unit_id).items()
            if qid == question_id
        )
        placeholder = qb_service.get_next_placeholder(unit_type, unit_id, q_type, is_ins)
        if not placeholder:
            return jsonify({'success': False, 'error': 'No available placeholders'})
    else:
        return jsonify({'success': False, 'error': 'Either select existing or provide new question details'})

    # Department admins need approval
    if current_user.role == 'dept_admin':
        pending_manager = get_pending_manager()
        # Get the question text for visual feedback (either new text or from existing question)
        display_text = question_text if question_text else qb_service.questions.get(question_id, {}).get('text', '')
        change_id = pending_manager.add_change(
            change_type='add',
            unit_type=unit_type,
            unit_id=unit_id,
            placeholder=placeholder,
            question_id=question_id,
            submitted_by=current_user.linkblue,
            college_code=college_code,
            new_text=display_text
        )

        log_audit('question_add_pending', current_user, {
            'unit_type': unit_type, 'unit_id': unit_id,
            'question_id': question_id, 'change_id': change_id
        })
        return jsonify({'success': True, 'pending': True, 'message': 'Change submitted for approval'})

    # Apply immediately for super/college admin
    # Create backup before change
    backup_service = get_backup_service()
    backup = backup_service.create_backup('qm', 'change', current_user,
                                           details={'unit_type': unit_type, 'unit_id': unit_id, 'action': 'add'})

    if qb_service.add_question_to_unit(unit_type, unit_id, placeholder, question_id):
        # Get question text for logging
        q_text = question_text if question_text else qb_service.questions.get(question_id, {}).get('text', '')
        log_audit('question_add', current_user, {
            'unit_type': unit_type, 'unit_id': unit_id,
            'placeholder': placeholder, 'question_id': question_id
        })
        # Log to database with question text for tracking visibility
        QBAuditLog.log_action('question_add', current_user,
                              details={'unit_type': unit_type, 'unit_id': unit_id,
                                       'placeholder': placeholder, 'question_id': question_id,
                                       'question_text': q_text[:150] if q_text else None},
                              backup_id=backup.id if backup else None)
        db.session.commit()
        return jsonify({'success': True})

    return jsonify({'success': False, 'error': 'Failed to add question'})


@questions_bp.route('/api/question/remove', methods=['POST'])
@api_qb_access_required
def api_remove_question():
    """API: Remove question from unit"""
    data = request.get_json()
    unit_type = data.get('unit_type')
    unit_id = data.get('unit_id')
    placeholder = data.get('placeholder')
    
    if not all([unit_type, unit_id, placeholder]):
        return jsonify({'success': False, 'error': 'Missing required fields'})
    
    qb_service = get_qb_service()

    # Ensure question bank is loaded to get question text
    if not qb_service.questions:
        qb_service.load_question_bank()

    current_mapping = qb_service._mapping_for_unit(unit_type, unit_id)
    question_id = current_mapping.get(placeholder, '')
    college_code = qb_service.get_college_for_unit(unit_id) or current_user.college_code

    # Get the question text for visual feedback
    question_text = qb_service.questions.get(question_id, {}).get('text', '')

    # Department admins need approval
    if current_user.role == 'dept_admin':
        pending_manager = get_pending_manager()
        change_id = pending_manager.add_change(
            change_type='remove',
            unit_type=unit_type,
            unit_id=unit_id,
            placeholder=placeholder,
            question_id=question_id,
            submitted_by=current_user.linkblue,
            college_code=college_code,
            old_value=question_text  # Store question text for visual feedback
        )
        
        log_audit('question_remove_pending', current_user, {
            'unit_type': unit_type, 'unit_id': unit_id,
            'question_id': question_id, 'change_id': change_id
        })
        return jsonify({'success': True, 'pending': True, 'message': 'Change submitted for approval'})
    
    # Apply immediately
    # Create backup before change
    backup_service = get_backup_service()
    backup = backup_service.create_backup('qm', 'change', current_user,
                                           details={'unit_type': unit_type, 'unit_id': unit_id, 'action': 'remove'})

    if qb_service.remove_question_from_unit(unit_type, unit_id, placeholder):
        log_audit('question_remove', current_user, {
            'unit_type': unit_type, 'unit_id': unit_id,
            'placeholder': placeholder, 'question_id': question_id
        })
        # Log to database with question text for tracking visibility
        QBAuditLog.log_action('question_remove', current_user,
                              details={'unit_type': unit_type, 'unit_id': unit_id,
                                       'placeholder': placeholder, 'question_id': question_id,
                                       'question_text': question_text[:150] if question_text else None},
                              backup_id=backup.id if backup else None)
        db.session.commit()
        return jsonify({'success': True})

    return jsonify({'success': False, 'error': 'Failed to remove question'})


@questions_bp.route('/pending')
@login_required
def pending_changes():
    """View pending changes for approval"""
    from app.models.admin import Admin
    from app.models.course import Course, Department, College

    if not (current_user.is_super_admin() or current_user.role == 'college_admin'):
        flash('You do not have permission to view pending changes.', 'danger')
        return redirect(url_for('questions.browser'))

    pending_manager = get_pending_manager()
    qb_service = get_qb_service()

    # Ensure question bank is loaded
    if not qb_service.questions:
        qb_service.load_question_bank()

    if current_user.is_super_admin():
        pending = pending_manager.get_pending()
    else:
        pending = pending_manager.get_pending(current_user.college_code)

    # Enrich pending changes with additional data
    enriched_pending = []
    for change in pending:
        enriched_change = dict(change)

        # Look up submitter's full name
        submitter = Admin.query.filter_by(linkblue=change.get('submitted_by')).first()
        if submitter:
            enriched_change['submitter_name'] = f"{submitter.first_name} {submitter.last_name}"
        else:
            enriched_change['submitter_name'] = change.get('submitted_by', 'Unknown')

        # Look up question text
        question_id = change.get('question_id')
        if question_id and question_id in qb_service.questions:
            enriched_change['question_text'] = qb_service.questions[question_id].get('text', '')
        else:
            enriched_change['question_text'] = ''

        # Determine hierarchy info for grouping
        unit_type = change.get('unit_type', '').upper()
        unit_id = change.get('unit_id', '')
        college_code = change.get('college_code', '')

        # Build hierarchy path
        enriched_change['college_code'] = college_code
        enriched_change['college_name'] = ''
        enriched_change['department_name'] = ''
        enriched_change['course_code'] = ''
        enriched_change['section_id'] = ''

        if college_code:
            college = College.query.filter_by(code=college_code).first()
            if college:
                enriched_change['college_name'] = college.name

        if unit_type == 'DEPARTMENT':
            dept = Department.query.filter_by(id=unit_id).first()
            if dept:
                enriched_change['department_name'] = dept.name
                if not college_code:
                    enriched_change['college_code'] = dept.college_code
                    college = College.query.filter_by(code=dept.college_code).first()
                    if college:
                        enriched_change['college_name'] = college.name
        elif unit_type == 'COURSE':
            enriched_change['course_code'] = unit_id
            # Try to find the course to get its department and college
            course = Course.query.filter_by(class_code=unit_id).first()
            if course:
                enriched_change['department_name'] = course.department.name if course.department else course.department_id
                if not college_code:
                    enriched_change['college_code'] = course.college_code
                    if course.college:
                        enriched_change['college_name'] = course.college.name
        elif unit_type == 'SECTION':
            enriched_change['section_id'] = unit_id
            course = Course.query.filter_by(section_key=unit_id).first()
            if course:
                enriched_change['course_code'] = course.class_code
                enriched_change['department_name'] = course.department.name if course.department else course.department_id
                if not college_code:
                    enriched_change['college_code'] = course.college_code
                    if course.college:
                        enriched_change['college_name'] = course.college.name

        enriched_pending.append(enriched_change)

    # Group the pending changes by hierarchy
    grouped_changes = {}
    for change in enriched_pending:
        if current_user.is_super_admin():
            # Group by College -> Department -> Course -> Section
            college_key = change.get('college_name') or change.get('college_code') or 'Unknown College'
        else:
            # For college admins, skip college grouping and start with department
            college_key = '__skip__'

        dept_key = change.get('department_name') or 'Unknown Department'
        course_key = change.get('course_code') or 'N/A'
        section_key = change.get('section_id') or 'N/A'

        if college_key not in grouped_changes:
            grouped_changes[college_key] = {}
        if dept_key not in grouped_changes[college_key]:
            grouped_changes[college_key][dept_key] = {}
        if course_key not in grouped_changes[college_key][dept_key]:
            grouped_changes[college_key][dept_key][course_key] = {}
        if section_key not in grouped_changes[college_key][dept_key][course_key]:
            grouped_changes[college_key][dept_key][course_key][section_key] = []

        grouped_changes[college_key][dept_key][course_key][section_key].append(change)

    return render_template('questions/pending.html',
                           pending_changes=enriched_pending,
                           grouped_changes=grouped_changes,
                           is_super_admin=current_user.is_super_admin())


@questions_bp.route('/pending/mine')
@login_required
def my_pending_changes():
    """View pending changes submitted by the current user"""
    pending_manager = get_pending_manager()
    pending = pending_manager.get_pending_for_submitter(current_user.linkblue)
    return render_template('questions/pending_mine.html', pending_changes=pending)


@questions_bp.route('/api/pending/approve', methods=['POST'])
@api_login_required
def api_approve_change():
    """API: Approve pending change"""
    if not (current_user.is_super_admin() or current_user.role == 'college_admin'):
        return jsonify({'success': False, 'error': 'Permission denied'})
    
    data = request.get_json()
    change_id = data.get('change_id')
    
    pending_manager = get_pending_manager()
    change = pending_manager.approve(change_id, current_user.linkblue)
    
    if change:
        qb_service = get_qb_service()

        # Create backup before applying change
        backup_service = get_backup_service()
        backup_type = 'qb' if change['type'] == 'edit' else 'qm'
        backup = backup_service.create_backup(backup_type, 'change', current_user,
                                               details={'change_id': change_id, 'change_type': change['type']})

        if change['type'] == 'add':
            # Get next available placeholder instead of using submitted one
            # This fixes the bug where multiple questions submitted with same placeholder
            # would overwrite each other when approved
            q = qb_service.questions.get(change['question_id'], {})
            q_type = q.get('type', 'Selection')
            # Determine if instructor question from placeholder naming convention
            submitted_placeholder = change.get('placeholder', '')
            is_instructor = 'ins_' in submitted_placeholder.lower() or '_ins_' in submitted_placeholder.lower()

            # Get the next available placeholder for this unit
            next_placeholder = qb_service.get_next_placeholder(
                change['unit_type'], change['unit_id'], q_type, is_instructor
            )

            if not next_placeholder:
                return jsonify({'success': False, 'error': 'No available placeholders for this unit'})

            qb_service.add_question_to_unit(change['unit_type'], change['unit_id'], next_placeholder, change['question_id'])
        elif change['type'] == 'remove':
            qb_service.remove_question_from_unit(change['unit_type'], change['unit_id'], change['placeholder'])
        elif change['type'] == 'edit':
            qb_service.update_question_details(
                change['question_id'],
                new_text=change.get('new_text'),
                new_type=change.get('new_type')
            )

        log_audit('change_approved', current_user, {'change_id': change_id, 'submitted_by': change['submitted_by']})

        # Log to database
        QBAuditLog.log_action('change_approved', current_user,
                              details={'change_id': change_id, 'change_type': change['type'],
                                       'submitted_by': change['submitted_by']},
                              backup_id=backup.id if backup else None)
        db.session.commit()

        return jsonify({'success': True})
    
    return jsonify({'success': False, 'error': 'Change not found'})


@questions_bp.route('/api/pending/reject', methods=['POST'])
@api_login_required
def api_reject_change():
    """API: Reject pending change"""
    if not (current_user.is_super_admin() or current_user.role == 'college_admin'):
        return jsonify({'success': False, 'error': 'Permission denied'})
    
    data = request.get_json()
    change_id = data.get('change_id')
    reason = data.get('reason', '')
    
    pending_manager = get_pending_manager()
    change = pending_manager.reject(change_id, current_user.linkblue, reason)
    
    if change:
        log_audit('change_rejected', current_user, {'change_id': change_id, 'reason': reason})

        # Log to database
        QBAuditLog.log_action('change_rejected', current_user,
                              details={'change_id': change_id, 'change_type': change.get('type'),
                                       'submitted_by': change.get('submitted_by'), 'reason': reason})
        db.session.commit()

        return jsonify({'success': True})

    return jsonify({'success': False, 'error': 'Change not found'})


@questions_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_files():
    """Import QB and QM files - Super Admin only"""
    if not current_user.is_super_admin():
        flash('Only super administrators can import files.', 'danger')
        return redirect(url_for('questions.browser'))

    if request.method == 'POST':
        os.makedirs(DATASOURCES_PATH, exist_ok=True)
        imported = []
        backup_service = get_backup_service()

        if 'qb_file' in request.files:
            qb_file = request.files['qb_file']
            if qb_file.filename:
                # Create backup before import
                backup = backup_service.create_backup('qb', 'import', current_user,
                                                       details={'original_filename': qb_file.filename})
                qb_file.save(os.path.join(DATASOURCES_PATH, QB_FILENAME))
                imported.append('Question Bank')
                log_audit('import_qb', current_user, {'filename': qb_file.filename})
                # Log to database
                QBAuditLog.log_action('qb_import', current_user,
                                      details={'filename': qb_file.filename},
                                      backup_id=backup.id if backup else None)

        if 'qm_file' in request.files:
            qm_file = request.files['qm_file']
            if qm_file.filename:
                # Create backup before import
                backup = backup_service.create_backup('qm', 'import', current_user,
                                                       details={'original_filename': qm_file.filename})
                qm_file.save(os.path.join(DATASOURCES_PATH, QM_FILENAME))
                imported.append('Question Mapping')
                log_audit('import_qm', current_user, {'filename': qm_file.filename})
                # Log to database
                QBAuditLog.log_action('qm_import', current_user,
                                      details={'filename': qm_file.filename},
                                      backup_id=backup.id if backup else None)

        if imported:
            db.session.commit()
            get_qb_service(force_reload=True)
            flash(f'Successfully imported: {", ".join(imported)}. Backups created automatically.', 'success')
        else:
            flash('No files were uploaded.', 'warning')

        return redirect(url_for('questions.browser'))
    
    # GET - show form
    current_files = []
    
    for fname, sheet in [(QB_FILENAME, 'Question Bank Questions'), (QM_FILENAME, None)]:
        fpath = os.path.join(DATASOURCES_PATH, fname)
        if os.path.exists(fpath):
            try:
                xlsx = pd.ExcelFile(fpath)
                sname = sheet if sheet else xlsx.sheet_names[0]
                df = pd.read_excel(xlsx, sheet_name=sname, header=None if fname == QM_FILENAME else 0)
                records = len(df) - 2 if fname == QM_FILENAME else len(df)
                current_files.append({
                    'name': fname, 'exists': True, 'records': records,
                    'modified': datetime.fromtimestamp(os.path.getmtime(fpath)).strftime('%Y-%m-%d %H:%M')
                })
            except:
                current_files.append({'name': fname, 'exists': True, 'records': '?', 'modified': '?'})
        else:
            current_files.append({'name': fname, 'exists': False})
    
    return render_template('questions/import.html', current_files=current_files)


@questions_bp.route('/export/qb')
@login_required
def export_qb():
    """Export QB - Super Admin only"""
    if not current_user.is_super_admin():
        flash('Only super administrators can export files.', 'danger')
        return redirect(url_for('questions.browser'))

    qb_service = get_qb_service()
    if not qb_service.questions:
        qb_service.load_question_bank()

    try:
        # Create backup on export
        backup_service = get_backup_service()
        backup = backup_service.create_backup('qb', 'export', current_user)

        output = qb_service.export_question_bank()
        log_audit('export_qb', current_user, {})

        # Log to database
        QBAuditLog.log_action('qb_export', current_user,
                              details={'question_count': len(qb_service.questions)},
                              backup_id=backup.id if backup else None)
        db.session.commit()

        return Response(
            output.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=QB_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
        )
    except Exception as e:
        flash(f'Export failed: {str(e)}', 'danger')
        return redirect(url_for('questions.browser'))


@questions_bp.route('/export/summary')
@qb_access_required
def export_summary():
    """Export a readable question assignment summary for college and super admins."""
    if current_user.role not in ('super_admin', 'college_admin'):
        flash('Only college and super administrators can export the question bank summary.', 'danger')
        return redirect(url_for('questions.browser'))
    if current_user.role == 'college_admin' and not current_user.college_code:
        flash('Your account must be assigned to a college before exporting a summary.', 'danger')
        return redirect(url_for('questions.browser'))

    college_code = None if current_user.is_super_admin() else current_user.college_code
    admin_scope = {'college': college_code, 'department': None} if college_code else None
    qb_service = get_qb_service()

    try:
        qb_service.load_courses(admin_scope=admin_scope)
        qb_service.load_question_bank()
        qb_service.load_question_mapping()
        output, summary = qb_service.export_question_bank_summary(college_code=college_code)

        audit_details = {
            'scope': college_code or 'all_colleges',
            'question_count': summary['row_count'],
            'sheet_count': summary['sheet_count'],
        }
        log_audit('export_qb_summary', current_user, audit_details)
        QBAuditLog.log_action('qb_summary_export', current_user, details=audit_details)
        db.session.commit()

        scope_label = college_code or 'all_colleges'
        filename = f'QB_summary_{scope_label}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return Response(
            output.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        db.session.rollback()
        flash(f'Summary export failed: {str(e)}', 'danger')
        return redirect(url_for('questions.browser'))


@questions_bp.route('/export/qm')
@login_required
def export_qm():
    """Export QM - Super Admin only"""
    if not current_user.is_super_admin():
        flash('Only super administrators can export files.', 'danger')
        return redirect(url_for('questions.browser'))

    qb_service = get_qb_service()
    if not qb_service.question_mapping:
        qb_service.load_question_mapping()

    try:
        # Create backup on export
        backup_service = get_backup_service()
        backup = backup_service.create_backup('qm', 'export', current_user)

        output = qb_service.export_question_mapping()
        log_audit('export_qm', current_user, {})

        # Log to database
        mapping_count = sum(len(m) for m in qb_service.question_mapping.values())
        QBAuditLog.log_action('qm_export', current_user,
                              details={'mapping_count': mapping_count},
                              backup_id=backup.id if backup else None)
        db.session.commit()

        return Response(
            output.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=QM_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
        )
    except Exception as e:
        flash(f'Export failed: {str(e)}', 'danger')
        return redirect(url_for('questions.browser'))
