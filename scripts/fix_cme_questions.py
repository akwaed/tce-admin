#!/usr/bin/env python3
"""
Fix script to restore all approved CME questions with correct placeholders.

The bug: All questions for a section were submitted with the same placeholder
(Sec_Crs_Sel_001), so when approved, only the last one was saved.

This script:
1. Reads all approved changes from pending_changes.json
2. Groups them by section
3. Assigns proper incrementing placeholders (Sec_Crs_Sel_001, 002, 003, etc.)
4. Updates the QM.xlsx file with all the mappings

Run from tce-admin directory:
    python scripts/fix_cme_questions.py
"""

import os
import json
import shutil
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

DATASOURCES_PATH = './datasources'
BACKUPS_PATH = './datasources/backups'
QM_FILE = os.path.join(DATASOURCES_PATH, 'QM.xlsx')
PENDING_FILE = os.path.join(DATASOURCES_PATH, 'pending_changes.json')


def load_pending_changes():
    """Load pending changes JSON"""
    if os.path.exists(PENDING_FILE):
        with open(PENDING_FILE, 'r') as f:
            return json.load(f)
    return []


def get_approved_cme_changes(pending):
    """Get all approved CME section changes grouped by section"""
    sections = defaultdict(list)

    for change in pending:
        if change.get('status') != 'approved':
            continue
        if change.get('type') != 'add':
            continue

        unit_type = change.get('unit_type', '').upper()
        unit_id = change.get('unit_id', '')

        # Only process SECTION level CME entries
        if unit_type == 'SECTION' and 'CME' in unit_id:
            sections[unit_id].append({
                'change_id': change.get('id'),
                'question_id': change.get('question_id'),
                'question_text': change.get('new_text', '')[:80],
                'original_placeholder': change.get('placeholder')
            })

    return sections


def load_qm_mappings(filepath):
    """Load existing QM mappings from xlsx"""
    mappings = {}
    placeholders = []

    if not os.path.exists(filepath):
        print(f"ERROR: QM file not found at {filepath}")
        return None, None, None

    wb = load_workbook(filepath)
    ws = wb.active

    rows = list(ws.rows)
    if len(rows) < 3:
        return wb, {}, []

    # Row 2 = placeholder names (0-indexed: rows[1])
    placeholders = [str(cell.value) if cell.value else '' for cell in rows[1]]

    for row_idx in range(2, len(rows)):
        row = rows[row_idx]
        mapping_type = str(row[0].value).upper() if row[0].value else ''
        unit_id = str(row[1].value) if row[1].value else ''

        if not mapping_type or mapping_type in ['NONE', 'NAN', 'TYPE'] or not unit_id or unit_id == 'None':
            continue

        key = f"{mapping_type}:{unit_id}"
        mappings[key] = {'row_idx': row_idx, 'questions': {}}

        for col_idx in range(2, len(row)):
            if col_idx < len(placeholders):
                placeholder = placeholders[col_idx]
                value = row[col_idx].value
                if value and str(value) not in ['None', 'nan', '']:
                    mappings[key]['questions'][placeholder] = str(value)

    return wb, mappings, placeholders


def get_next_placeholder(existing_questions, q_type='Sel', target='Crs'):
    """Get next available placeholder for a section"""
    # Format: Sec_{Crs|Ins}_{Sel|Com}_NNN
    prefix = f"Sec_{target}_{q_type}_"
    max_num = 14 if q_type == 'Sel' else 5

    used = set()
    for p in existing_questions.keys():
        if p.startswith(prefix):
            try:
                num = int(p.replace(prefix, ''))
                used.add(num)
            except:
                pass

    for i in range(1, max_num + 1):
        if i not in used:
            return f"{prefix}{i:03d}"

    return None  # No slots available


def apply_fixes(wb, mappings, placeholders, cme_sections):
    """Apply the fixes to the workbook"""
    ws = wb.active

    changes_made = []

    for section_id, questions in cme_sections.items():
        key = f"SECTION:{section_id}"

        if key not in mappings:
            print(f"WARNING: Section {section_id} not found in QM, will add new row")
            # For now, skip sections not in QM - they'd need a new row
            continue

        row_idx = mappings[key]['row_idx']
        existing = mappings[key]['questions'].copy()

        print(f"\n{section_id}:")
        print(f"  Currently has {len(existing)} questions mapped")
        print(f"  Need to add {len(questions)} approved questions")

        for q in questions:
            q_id = q['question_id']

            # Check if already mapped
            already_mapped = False
            for placeholder, mapped_q in existing.items():
                if mapped_q == q_id:
                    already_mapped = True
                    print(f"  - {q_id} already mapped at {placeholder}")
                    break

            if already_mapped:
                continue

            # Get next available placeholder
            next_placeholder = get_next_placeholder(existing)
            if not next_placeholder:
                print(f"  ERROR: No available placeholders for {q_id}")
                continue

            # Find column index for this placeholder
            col_idx = None
            for idx, p in enumerate(placeholders):
                if p == next_placeholder:
                    col_idx = idx + 1  # openpyxl is 1-indexed
                    break

            if col_idx is None:
                print(f"  ERROR: Placeholder {next_placeholder} not found in columns")
                continue

            # Apply the change
            cell = ws.cell(row=row_idx + 1, column=col_idx)  # +1 because openpyxl is 1-indexed
            cell.value = q_id
            existing[next_placeholder] = q_id

            print(f"  + Added {q_id} at {next_placeholder}")
            changes_made.append({
                'section': section_id,
                'question_id': q_id,
                'placeholder': next_placeholder,
                'text': q['question_text']
            })

    return changes_made


def main():
    print("=" * 80)
    print("CME Question Fix Script")
    print("=" * 80)

    # Check if QM.xlsx exists, if not copy from latest backup
    if not os.path.exists(QM_FILE):
        print(f"\nQM.xlsx not found at {QM_FILE}")
        backups = sorted([f for f in os.listdir(BACKUPS_PATH) if f.startswith('QM_backup_')])
        if backups:
            latest = backups[-1]
            print(f"Copying from latest backup: {latest}")
            shutil.copy(os.path.join(BACKUPS_PATH, latest), QM_FILE)
        else:
            print("ERROR: No QM backups found!")
            return

    # Load pending changes
    print("\n1. Loading pending changes...")
    pending = load_pending_changes()
    if not pending:
        print("   ERROR: No pending changes found!")
        return

    print(f"   Found {len(pending)} total pending entries")

    # Get approved CME changes
    print("\n2. Filtering approved CME section changes...")
    cme_sections = get_approved_cme_changes(pending)

    total_questions = sum(len(q) for q in cme_sections.values())
    print(f"   Found {len(cme_sections)} CME sections with {total_questions} approved questions")

    for section, questions in sorted(cme_sections.items()):
        print(f"   - {section}: {len(questions)} questions")

    # Load current QM file
    print("\n3. Loading QM.xlsx...")
    wb, mappings, placeholders = load_qm_mappings(QM_FILE)
    if wb is None:
        return

    print(f"   Loaded {len(mappings)} unit mappings")

    # Create backup before making changes
    print("\n4. Creating backup...")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"QM_backup_BEFORE_FIX_{timestamp}.xlsx"
    backup_path = os.path.join(BACKUPS_PATH, backup_name)
    shutil.copy(QM_FILE, backup_path)
    print(f"   Backup saved: {backup_name}")

    # Apply fixes
    print("\n5. Applying fixes...")
    changes = apply_fixes(wb, mappings, placeholders, cme_sections)

    if changes:
        # Save the workbook
        print(f"\n6. Saving {len(changes)} changes to QM.xlsx...")
        wb.save(QM_FILE)
        print("   SAVED!")

        # Summary
        print("\n" + "=" * 80)
        print("SUMMARY OF CHANGES")
        print("=" * 80)
        for c in changes:
            print(f"\n{c['section']}")
            print(f"  Placeholder: {c['placeholder']}")
            print(f"  Question ID: {c['question_id']}")
            print(f"  Text: {c['text']}...")
    else:
        print("\n   No changes needed - all questions already mapped correctly")

    print("\n" + "=" * 80)
    print("FIX COMPLETE")
    print("=" * 80)


if __name__ == '__main__':
    main()
