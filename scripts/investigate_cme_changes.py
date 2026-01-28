#!/usr/bin/env python3
"""
Diagnostic script to investigate CME/che202 question changes
Run this from the tce-admin directory:
    python scripts/investigate_cme_changes.py
"""

import os
import json
from datetime import datetime
from openpyxl import load_workbook

DATASOURCES_PATH = './datasources'
BACKUPS_PATH = './datasources/backups'

def load_pending_changes():
    """Load pending changes JSON if it exists"""
    path = os.path.join(DATASOURCES_PATH, 'pending_changes.json')
    if os.path.exists(path):
        with open(path, 'r') as f:
            return json.load(f)
    return []

def load_audit_log():
    """Load JSON audit log if it exists"""
    path = os.path.join(DATASOURCES_PATH, 'qb_audit_log.json')
    if os.path.exists(path):
        with open(path, 'r') as f:
            return json.load(f)
    return []

def get_qm_backups():
    """Get list of QM backups sorted by timestamp"""
    backups = []
    if os.path.exists(BACKUPS_PATH):
        for f in os.listdir(BACKUPS_PATH):
            if f.startswith('QM_backup_') and f.endswith('.xlsx'):
                try:
                    ts_str = f.replace('QM_backup_', '').replace('.xlsx', '')
                    ts = datetime.strptime(ts_str, '%Y%m%d_%H%M%S')
                    backups.append({'filename': f, 'timestamp': ts})
                except:
                    pass
    return sorted(backups, key=lambda x: x['timestamp'])

def load_qm_file(filepath):
    """Load question mapping from QM xlsx file"""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.rows)
        if len(rows) < 3:
            return {}

        # Row 1 = UUIDs, Row 2 = placeholder names (0-indexed: rows[0], rows[1])
        placeholders = [str(cell.value) if cell.value else '' for cell in rows[1]]

        mappings = {}
        for row_idx in range(2, len(rows)):
            row = rows[row_idx]
            mapping_type = str(row[0].value).upper() if row[0].value else ''
            unit_id = str(row[1].value) if row[1].value else ''

            if not mapping_type or mapping_type in ['NONE', 'NAN', 'TYPE'] or not unit_id or unit_id == 'None':
                continue

            key = f"{mapping_type}:{unit_id}"
            mappings[key] = {}
            for col_idx in range(2, len(row)):
                if col_idx < len(placeholders):
                    placeholder = placeholders[col_idx]
                    value = row[col_idx].value
                    if value and str(value) not in ['None', 'nan', '']:
                        mappings[key][placeholder] = str(value)

        wb.close()
        return mappings
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
        return {}

def load_qb_file(filepath):
    """Load question bank from QB xlsx file"""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        questions = {}

        if 'Question Bank Questions' in wb.sheetnames:
            ws = wb['Question Bank Questions']
            rows = list(ws.rows)

            # Find header row
            headers = {str(cell.value): idx for idx, cell in enumerate(rows[0])}

            q_id_col = headers.get('Question Id', 0)
            q_title_col = headers.get('Question Title', 1)

            for row in rows[1:]:
                q_id = str(row[q_id_col].value) if row[q_id_col].value else ''
                if q_id and q_id not in ['None', 'nan']:
                    q_title = str(row[q_title_col].value) if row[q_title_col].value else ''
                    questions[q_id] = {'text': q_title}

        wb.close()
        return questions
    except Exception as e:
        print(f"Error loading QB {filepath}: {e}")
    return {}

def compare_backups(backup1, backup2, qb_questions):
    """Compare two QM backups and show differences"""
    path1 = os.path.join(BACKUPS_PATH, backup1['filename'])
    path2 = os.path.join(BACKUPS_PATH, backup2['filename'])

    map1 = load_qm_file(path1)
    map2 = load_qm_file(path2)

    changes = []

    # Find additions (in map2 but not in map1)
    for key, questions in map2.items():
        old_questions = map1.get(key, {})
        for placeholder, q_id in questions.items():
            if placeholder not in old_questions or old_questions[placeholder] != q_id:
                q_text = qb_questions.get(q_id, {}).get('text', 'Unknown question')
                changes.append({
                    'type': 'ADDED',
                    'unit': key,
                    'placeholder': placeholder,
                    'question_id': q_id,
                    'question_text': q_text
                })

    # Find removals (in map1 but not in map2)
    for key, questions in map1.items():
        new_questions = map2.get(key, {})
        for placeholder, q_id in questions.items():
            if placeholder not in new_questions:
                q_text = qb_questions.get(q_id, {}).get('text', 'Unknown question')
                changes.append({
                    'type': 'REMOVED',
                    'unit': key,
                    'placeholder': placeholder,
                    'question_id': q_id,
                    'question_text': q_text
                })

    return changes

def main():
    print("=" * 80)
    print("CME/che202 Question Changes Investigation")
    print("=" * 80)

    # Check for pending changes
    print("\n1. PENDING CHANGES FILE:")
    pending = load_pending_changes()
    if pending:
        che202_changes = [c for c in pending if c.get('submitted_by') == 'che202']
        approved_changes = [c for c in pending if c.get('status') == 'approved']
        print(f"   Total pending entries: {len(pending)}")
        print(f"   From che202: {len(che202_changes)}")
        print(f"   Approved status: {len(approved_changes)}")

        # Show all che202 changes
        for c in pending:
            if c.get('submitted_by') == 'che202':
                print(f"\n   ID {c.get('id')}: {c.get('type')} - Status: {c.get('status')}")
                print(f"   Unit: {c.get('unit_type')}:{c.get('unit_id')}")
                print(f"   Placeholder: {c.get('placeholder')}")
                print(f"   Question ID: {c.get('question_id')}")
                if c.get('new_text'):
                    print(f"   Text: {c.get('new_text')[:80]}...")
                if c.get('old_value'):
                    print(f"   Old value: {c.get('old_value')[:80]}...")
    else:
        print("   No pending_changes.json found or file is empty")

    # Check audit log
    print("\n2. JSON AUDIT LOG:")
    audit = load_audit_log()
    if audit:
        che202_audit = [a for a in audit if a.get('user') == 'che202']
        print(f"   Total audit entries: {len(audit)}")
        print(f"   From che202: {len(che202_audit)}")
        for a in che202_audit[-10:]:  # Last 10
            print(f"   - {a.get('timestamp')}: {a.get('action')}")
            details = a.get('details', {})
            if details:
                for k, v in details.items():
                    print(f"     {k}: {v}")
    else:
        print("   No qb_audit_log.json found or file is empty")

    # Load QB for question texts
    print("\n3. LOADING QUESTION BANK...")
    qb_questions = {}
    qb_backups = sorted([f for f in os.listdir(BACKUPS_PATH) if f.startswith('QB_backup_')]) if os.path.exists(BACKUPS_PATH) else []
    if qb_backups:
        latest_qb = qb_backups[-1]
        qb_questions = load_qb_file(os.path.join(BACKUPS_PATH, latest_qb))
        print(f"   Loaded {len(qb_questions)} questions from {latest_qb}")
    elif os.path.exists(os.path.join(DATASOURCES_PATH, 'QB.xlsx')):
        qb_questions = load_qb_file(os.path.join(DATASOURCES_PATH, 'QB.xlsx'))
        print(f"   Loaded {len(qb_questions)} questions from QB.xlsx")
    else:
        print("   No QB file found!")

    # Analyze QM backups
    print("\n4. ANALYZING QM BACKUPS (Jan 27, 2026 changes):")
    qm_backups = get_qm_backups()

    # Filter to Jan 27 backups
    jan27_backups = [b for b in qm_backups if b['timestamp'].date() == datetime(2026, 1, 27).date()]
    print(f"   Found {len(jan27_backups)} QM backups from Jan 27, 2026")

    if len(jan27_backups) >= 2:
        print(f"\n   First backup: {jan27_backups[0]['filename']} ({jan27_backups[0]['timestamp']})")
        print(f"   Last backup: {jan27_backups[-1]['filename']} ({jan27_backups[-1]['timestamp']})")
        print("-" * 60)

        changes = compare_backups(jan27_backups[0], jan27_backups[-1], qb_questions)

        if changes:
            print(f"\n   TOTAL CHANGES ON JAN 27: {len(changes)}")
            added = [c for c in changes if c['type'] == 'ADDED']
            removed = [c for c in changes if c['type'] == 'REMOVED']
            print(f"   Added: {len(added)}, Removed: {len(removed)}")

            print("\n   ADDED QUESTIONS:")
            for c in added:
                print(f"\n   Unit: {c['unit']}")
                print(f"   Placeholder: {c['placeholder']}")
                print(f"   Question ID: {c['question_id']}")
                print(f"   Text: {c['question_text'][:100]}{'...' if len(c['question_text']) > 100 else ''}")

            if removed:
                print("\n   REMOVED QUESTIONS:")
                for c in removed:
                    print(f"\n   Unit: {c['unit']}")
                    print(f"   Placeholder: {c['placeholder']}")
                    print(f"   Question ID: {c['question_id']}")
        else:
            print("   No changes detected between first and last backup")

    # Show current state of DEPARTMENT mappings
    print("\n5. CURRENT DEPARTMENT MAPPINGS (from latest backup):")
    if qm_backups:
        latest_qm = qm_backups[-1]['filename']
        mappings = load_qm_file(os.path.join(BACKUPS_PATH, latest_qm))

        dept_mappings = {k: v for k, v in mappings.items() if k.startswith('DEPARTMENT:')}
        print(f"   Found {len(dept_mappings)} department mappings")

        for key, questions in sorted(dept_mappings.items()):
            if questions:
                print(f"\n   {key} ({len(questions)} questions):")
                for placeholder, q_id in sorted(questions.items()):
                    q_text = qb_questions.get(q_id, {}).get('text', 'Unknown')
                    print(f"     {placeholder}: [{q_id}] {q_text[:50]}...")

    print("\n" + "=" * 80)
    print("INVESTIGATION COMPLETE")
    print("=" * 80)

if __name__ == '__main__':
    main()
